from openpyxl import load_workbook
import os

nameFile=str(input("Введите имя файла в формате *.xlsx: ")) # имя файла с которым нужно работать
pass_file = os.path.abspath(nameFile)
wb_otchet = load_workbook(pass_file)
ws_otchet = wb_otchet.active
ws_otchet.title = "Исходный отчет"


def dellPristavla(spisok, separator):
    # функция принемает список с разными номерами, и по сепаратару отделяет нужную чвсть номера/данных и возвращает список с этими данными
    for i in range(0, len(spisok)):
        newData = spisok[i].split(separator)
        try:
            spisok[i] = newData[1] # тут может быть ошибка если в списке не будет нужного индекса
        except IndexError:
            spisok[i] = " "
    return spisok


nameTech = []
gosNumber = []
hozNumber = []
zapas = []

for i in range(11, ws_otchet.max_row+1):
    # print(ws_otchet.cell(row=i, column=1, ).value)
    stroka = ws_otchet.cell(row=i, column=1, ).value

    strokaList = stroka.split(", ")
    nameTech.append(strokaList[0])
    try:
        gosNumber.append(strokaList[1])
    except IndexError:
        gosNumber.append(" ")
    try:
        hozNumber.append(strokaList[2])
    except IndexError:
        hozNumber.append(" ")
    try:
        zapas.append(strokaList[3])
        zapas.append(ws_otchet.cell(row=i, column=6, ).value)
    except IndexError:
        zapas.append(" ")
gosNumber = dellPristavla(gosNumber, ".№")
hozNumber = dellPristavla(hozNumber, "Хоз.№")

new_ws_otchet = wb_otchet.create_sheet("Преобразованные данные")
new_ws_otchet.cell(row=1, column=1, value="Техника")
new_ws_otchet.cell(row=1, column=2, value="Гос. номер")
new_ws_otchet.cell(row=1, column=3, value="Хоз. номер")
for i in range(0, len(nameTech)):    
    new_ws_otchet.cell(row=i+2, column=1, value=nameTech[i])
    new_ws_otchet.cell(row=i+2, column=2, value=gosNumber[i])
    new_ws_otchet.cell(row=i+2, column=3, value=hozNumber[i])
    new_ws_otchet.cell(row=i+2, column=4, value=zapas[i])

wb_otchet.save(pass_file)
print(f"Проверяйте файлик {nameFile} ;)")
